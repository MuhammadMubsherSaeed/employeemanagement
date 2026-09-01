from __future__ import annotations

import csv
import io
from collections.abc import Callable, Sequence

from django.http import FileResponse, StreamingHttpResponse
from django.utils import timezone
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework.exceptions import ValidationError

from apps.attendance.services import CompanyClock
from apps.common.events import emit
from apps.common.tenancy import get_tenant_context
from apps.companies.services import get_company_settings

PDF_MAX_ROWS = 2000
EXPORT_FORMATS = frozenset({"csv", "xlsx", "pdf"})
_SAFE_FILTER_KEYS = (
    "date_from",
    "date_to",
    "employee",
    "department",
    "status",
    "employment_type",
    "leave_type",
    "search",
    "ordering",
)


class Echo:
    def write(self, value):
        return value


def resolve_export_format(raw: str | None) -> str:
    value = (raw or "csv").strip().lower()
    if value not in EXPORT_FORMATS:
        raise ValidationError({"format": ["Unsupported export format."]})
    return value


def safe_filter_metadata(params) -> dict[str, str]:
    metadata: dict[str, str] = {}
    for key in _SAFE_FILTER_KEYS:
        value = params.get(key)
        if value not in (None, ""):
            metadata[key] = str(value)
    return metadata


def company_report_date(request):
    ctx = get_tenant_context(request)
    if ctx.company is None:
        return timezone.now().date()
    settings = get_company_settings(ctx.company)
    return CompanyClock(settings).local_date()


def iter_rows(queryset, *, prefetch: bool):
    if prefetch:
        offset = 0
        size = 200
        while True:
            chunk = list(queryset[offset : offset + size])
            if not chunk:
                break
            yield from chunk
            if len(chunk) < size:
                break
            offset += size
        return
    yield from queryset.iterator(chunk_size=500)


class ExportService:
    def export(
        self,
        *,
        request,
        queryset,
        title: str,
        filename_stem: str,
        columns: Sequence[tuple[str, Callable]],
        export_format: str,
        report_type: str,
        prefetch: bool = False,
    ):
        ctx = get_tenant_context(request)
        export_format = resolve_export_format(export_format)
        if export_format == "pdf" and queryset.count() > PDF_MAX_ROWS:
            raise ValidationError(
                {
                    "format": [
                        "PDF export is limited to "
                        f"{PDF_MAX_ROWS} rows. Narrow the filters or use CSV/Excel."
                    ]
                }
            )
        stamp = company_report_date(request).isoformat()
        filename = f"{filename_stem}-{stamp}.{export_format}"
        headers = [name for name, _ in columns]
        filters = safe_filter_metadata(request.query_params)
        generated = timezone.now().isoformat()

        emit(
            "report.exported",
            actor=request.user,
            company=ctx.company,
            resource="reports.Report",
            resource_id=report_type,
            metadata={"format": export_format, "filters": filters},
        )

        if export_format == "csv":
            return self._csv(filename, headers, queryset, columns, prefetch)
        if export_format == "xlsx":
            return self._xlsx(
                filename, title, generated, headers, queryset, columns, prefetch
            )
        return self._pdf(
            filename,
            title,
            generated,
            filters,
            headers,
            queryset,
            columns,
            prefetch,
        )

    def _csv(self, filename, headers, queryset, columns, prefetch):
        def rows():
            writer = csv.writer(Echo())
            yield writer.writerow(headers)
            for instance in iter_rows(queryset, prefetch=prefetch):
                yield writer.writerow(
                    [_cell(getter(instance)) for _, getter in columns]
                )

        response = StreamingHttpResponse(rows(), content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _xlsx(self, filename, title, generated, headers, queryset, columns, prefetch):
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet(title="Report")
        sheet.append([title])
        sheet.append([f"Generated {generated}"])
        sheet.append([])
        sheet.append(headers)
        for instance in iter_rows(queryset, prefetch=prefetch):
            sheet.append([_cell(getter(instance)) for _, getter in columns])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = FileResponse(
            buffer,
            as_attachment=True,
            filename=filename,
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        return response

    def _pdf(
        self,
        filename,
        title,
        generated,
        filters,
        headers,
        queryset,
        columns,
        prefetch,
    ):
        rows: list[list[str]] = [headers]
        for instance in iter_rows(queryset, prefetch=prefetch):
            rows.append([_cell(getter(instance)) for _, getter in columns])
        buffer = io.BytesIO()
        document = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=24,
            rightMargin=24,
            topMargin=24,
            bottomMargin=24,
            title=title,
        )
        styles = getSampleStyleSheet()
        story = [
            Paragraph(title, styles["Title"]),
            Paragraph(f"Generated {generated}", styles["Normal"]),
        ]
        if filters:
            applied = ", ".join(f"{key}={value}" for key, value in filters.items())
            story.append(Paragraph(f"Filters: {applied}", styles["Normal"]))
        story.append(Spacer(1, 12))
        table = Table(rows, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(table)
        document.build(story)
        buffer.seek(0)
        return FileResponse(
            buffer,
            as_attachment=True,
            filename=filename,
            content_type="application/pdf",
        )


def _cell(value) -> str:
    if value is None:
        return ""
    return str(value)
